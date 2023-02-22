import streamlit as st
import pandas as pd
import numpy as np
import datetime
from datetime import date
import os
from os.path import exists
import plotly.express as px
import plotly.graph_objects as go
from PIL import Image
from st_clickable_images import clickable_images
from openpyxl import load_workbook
import io
from io import BytesIO
import xlsxwriter


# '''
# MAIN TODOs
# - select tabs per wkbk for a given offer - output on corresponding sheet
# - lookup cat in cattype
#
# '''

promo = "FLR"
row_ct = {}

balloons = False
clicked = 'https://images-wixmp-ed30a86b8c4ca887773594c2.wixmp.com/f/d3789c8c-0874-407c-a457-03b147f59b18/der4qbq-d8305001-dc6d-441e-ab8c-132b6f35fe63.png/v1/fill/w_1176,h_627,strp/lighting_mcqueen_by_darkmoonanimation_der4qbq-fullview.png?token=eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJzdWIiOiJ1cm46YXBwOjdlMGQxODg5ODIyNjQzNzNhNWYwZDQxNWVhMGQyNmUwIiwiaXNzIjoidXJuOmFwcDo3ZTBkMTg4OTgyMjY0MzczYTVmMGQ0MTVlYTBkMjZlMCIsIm9iaiI6W1t7ImhlaWdodCI6Ijw9NjI3IiwicGF0aCI6IlwvZlwvZDM3ODljOGMtMDg3NC00MDdjLWE0NTctMDNiMTQ3ZjU5YjE4XC9kZXI0cWJxLWQ4MzA1MDAxLWRjNmQtNDQxZS1hYjhjLTEzMmI2ZjM1ZmU2My5wbmciLCJ3aWR0aCI6Ijw9MTE3NiJ9XV0sImF1ZCI6WyJ1cm46c2VydmljZTppbWFnZS5vcGVyYXRpb25zIl19.1kNCLY8PHLdG7veIX1SA9cSn8HgDT0DeYvtJUSVNXTo'

# Convert cell range to pandas df
def range_to_df(ws, remove_nan=True):
    # Read the cell values into a list of lists
    data_rows = []
    for row in ws:
        data_cols = []
        for cell in row:
            data_cols.append(cell.value)
        data_rows.append(data_cols)
    df = pd.DataFrame(data_rows[1:])
    df.columns = data_rows[0]
    if remove_nan:
        df.dropna(axis=1, how='all', inplace=True)
    return df

def balloons():
    balloons = True

# Layout
st.set_page_config(layout="wide")
st.title('Fenced Offers - Upload and KACHOW!!!')
passwd = st.text_input('Enter Password')
if passwd == 'lightning':
    st.markdown('We process **faster** than lightning!')
    st.warning("Please upload your Fenced Offer input worksheet below.")


    # Excel INPUT
    uploaded_files = st.file_uploader("Select all ship workbooks (use ctrl to select multiple files)", accept_multiple_files=True, key='upload01')
    if len(uploaded_files) > 0:
        st.warning("Click on Lightning McQueen to run process.")
    # for uploaded_file in uploaded_files:
    #     bytes_data = uploaded_file.read()
    #     st.write("filename:", uploaded_file.name)
    #     st.write(bytes_data)

    # The car, the myth, the legend, KACHOW
    clicked = clickable_images(
        [
            'https://images-wixmp-ed30a86b8c4ca887773594c2.wixmp.com/f/d3789c8c-0874-407c-a457-03b147f59b18/der4qbq-d8305001-dc6d-441e-ab8c-132b6f35fe63.png/v1/fill/w_1176,h_627,strp/lighting_mcqueen_by_darkmoonanimation_der4qbq-fullview.png?token=eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJzdWIiOiJ1cm46YXBwOjdlMGQxODg5ODIyNjQzNzNhNWYwZDQxNWVhMGQyNmUwIiwiaXNzIjoidXJuOmFwcDo3ZTBkMTg4OTgyMjY0MzczYTVmMGQ0MTVlYTBkMjZlMCIsIm9iaiI6W1t7ImhlaWdodCI6Ijw9NjI3IiwicGF0aCI6IlwvZlwvZDM3ODljOGMtMDg3NC00MDdjLWE0NTctMDNiMTQ3ZjU5YjE4XC9kZXI0cWJxLWQ4MzA1MDAxLWRjNmQtNDQxZS1hYjhjLTEzMmI2ZjM1ZmU2My5wbmciLCJ3aWR0aCI6Ijw9MTE3NiJ9XV0sImF1ZCI6WyJ1cm46c2VydmljZTppbWFnZS5vcGVyYXRpb25zIl19.1kNCLY8PHLdG7veIX1SA9cSn8HgDT0DeYvtJUSVNXTo'
        ],
        titles=[f"Image #{str(i)}" for i in range(5)],
        div_style={"display": "flex", "justify-content": "center", "flex-wrap": "wrap"},
        img_style={"margin": "5px", "height": "200px"},
    )

    #===============================================================================
    # This is where the magic happens (when lightning is clicked) - KACHOW
    if clicked == 0:

        # Excel INPUT
        uploaded_file = 'Open FO Tracking.xlsx'
        sheet_name = 'Open FO Tracking'

        #TODO: form dictionary for lookup tables
        ship_dict = {'DD':'Disney Dream', 'DF':'Disney Fantasy', 'DW':'Disney Wonder', 'DM':'Disney Magic', 'WW':'Disney Wish'}


        # load lookup data
        lkup_wb = load_workbook(filename = 'LookupWKBK.xlsx', data_only=True)
        itens_ws = lkup_wb['Itineraries']
        lkup_itens = range_to_df(itens_ws['A2':'I750'])

        # ports
        lkup_ws = lkup_wb['Lookups']
        lkup_ports = range_to_df(lkup_ws['E2':'F300'])

        # taxes
        taxes_ws = lkup_wb['Taxes']
        lkup_taxes = range_to_df(taxes_ws['A1':'F500'])

        # cabins
        cabins_ws = lkup_wb['Cabins']
        lkup_cabins = range_to_df(cabins_ws['A1':'C500'])

        # Loop through submitted files
        # Load wkbk and sheet
        wb = load_workbook(filename = uploaded_file, data_only=True)

        # pull Pricing cache file
        # filename = "FO_Pricing_Cache.xlsx"
        # pricing_wb = load_workbook(filename = uploaded_file, data_only=True)
        # pricing_ws = pricing_wb['Cabins']
        # pricing_rg = range_to_df(pricing_ws['A4':'V500'])
        #
        # print(pricing_rg)

        #TODO: multiple sheets (one per offer, 'sail' in promoted fills in the lower table row 28, excluding G-J pricing)

        # print("-->", sheet_name, "\n")
        ws = wb[sheet_name]

        # Call pd convert function
        input = range_to_df(ws['A3':'AE100'])

        # Rename Columns to be unique (to combine and index)
        promo_list = ['GT','FLR','MTO','Interline','Cast','DCLCast','TAAP']
        cols = []
        promo_ct = -1
        for column in input.columns:
            if column in ('V','O','I','Cat.'):
                if column == "V":
                    promo_ct+=1
                cols.append(f'{column.replace(" ", "")}_{promo_list[promo_ct]}')
            else:
                cols.append(column)
        input.columns = cols

        input = input.loc[input['Ship'].notnull()]

        offers = input


        # Pull in template
        file_path = 'Template_MTO.xlsx' # generic template is called MTO
        tb = load_workbook(filename = file_path)
        ts = tb['OUTPUT']

        enter_row = 9

        # for colName in colNames:
        emptyRowCount = 200
        # offers = input[pd.to_numeric(input[colName], errors='coerce').notnull()]

        # check if offers is empty (will error on previous line and not update lastoffers)
        # lastoffers = sheet_name

        # filter to promo

        # offers = input.loc[input['V_'+promo].notnull()]

        # separate sail and non sail offers
        offers_sail = offers[offers['Cat._'+promo]=='Sail']
        offers = offers[offers['Cat._'+promo]!='Sail']

        # filter out empty Cat. rows
        offers = offers.loc[offers['Cat._'+promo].notnull()]
        offers_sail = offers_sail.loc[offers_sail['Cat._'+promo].notnull()]

        print(offers[['Ship', 'Sailing', 'V_'+promo, 'O_'+promo,'I_'+promo, 'Cat._'+promo]])

        # insert information
        for i in range(len(offers)):
            tdy = date.today()
            first_mon = tdy - datetime.timedelta(days = tdy.weekday())
            second_mon = tdy + datetime.timedelta(days = -tdy.weekday(), weeks=1)
            ts['F2'] = promo+" Sheet"
            ts['F4'] = "Rates Valid Monday, "+str(first_mon)+" through Monday, "+str(second_mon)

            ship = offers.iloc[i]['Ship']
            saildatefrom = offers.iloc[i]['Sailing']
            iten = lkup_itens[(lkup_itens['SHIP_CODE']==ship)&(lkup_itens['SAIL_DATE_FROM']==saildatefrom)]
            days = iten['DAYS'].values[0]
            cattype = offers.iloc[i]['Cat._'+promo]
            cattypetrans = lkup_cabins[(lkup_cabins['SHIP_CODE']==ship)&(lkup_cabins['CABIN_CATEGORY']==cattype)]['CAT_TYP'].values[0]
            vf_0 = int(offers.iloc[i][cattypetrans+'_'+promo])

            if "GT" in promo or "Cast" in promo:
                cattypeadd = "GT"
            else:
                cattypeadd = ''

            ts['B'+str(enter_row)] = ship_dict[ship]
            ts['C'+str(enter_row)] = saildatefrom
            ts['D'+str(enter_row)] = days
            ts['E'+str(enter_row)] = iten['FO Sheet Name'].values[0]
            # ts['E'+str(enter_row)] = iten['FO Sheet Name'].values[0].split(" ending")[0].split(" 2")[0]
            ts['F'+str(enter_row)] = lkup_ports[lkup_ports['Code'] == iten['PORT_FROM'].values[0]]['Name'].values[0].capitalize()
            ts['G'+str(enter_row)] = cattype+cattypeadd
            ts['H'+str(enter_row)] = vf_0
            ts['I'+str(enter_row)] = vf_0*days
            ts['J'+str(enter_row)] = 2*vf_0*days
            ts['K'+str(enter_row)] = lkup_taxes[(lkup_taxes['SHIP']==ship)&(lkup_taxes['SAIL_FROM']==saildatefrom)]['GVT_TAX'].values[0]

            # set new enter_row
            enter_row += 1


        # remove trailing empty rows (cleans up layout)
        ts.delete_rows(enter_row + 1, emptyRowCount - enter_row)

        # move down to next empty row
        enter_row+=3


        for i in range(len(offers_sail)):
            tdy = date.today()
            first_mon = tdy - datetime.timedelta(days = tdy.weekday())
            second_mon = tdy + datetime.timedelta(days = -tdy.weekday(), weeks=1)
            ts['F2'] = promo+" Sheet"
            ts['F4'] = "Rates Valid Monday, "+str(first_mon)+" through Monday, "+str(second_mon)

            ship = offers_sail.iloc[i]['Ship']
            saildatefrom = offers_sail.iloc[i]['Sailing']
            iten = lkup_itens[(lkup_itens['SHIP_CODE']==ship)&(lkup_itens['SAIL_DATE_FROM']==saildatefrom)]
            days = iten['DAYS'].values[0]
            # cattype = offers_sail.iloc[i]['Cat._'+promo]
            # cattypetrans = lkup_cabins[(lkup_cabins['SHIP_CODE']==ship)&(lkup_cabins['CABIN_CATEGORY']==cattype)]['CAT_TYP'].values[0]
            # vf_0 = int(offers_sail.iloc[i][cattypetrans+'_'+promo])

            if "GT" in promo or "Cast" in promo:
                cattypeadd = "GT"
            else:
                cattypeadd = ''

            ts['B'+str(enter_row)] = ship_dict[ship]
            ts['C'+str(enter_row)] = saildatefrom
            ts['D'+str(enter_row)] = days
            ts['E'+str(enter_row)] = iten['FO Sheet Name'].values[0]
            # ts['E'+str(enter_row)] = iten['FO Sheet Name'].values[0].split(" ending")[0].split(" 2")[0]
            ts['F'+str(enter_row)] = lkup_ports[lkup_ports['Code'] == iten['PORT_FROM'].values[0]]['Name'].values[0].capitalize()
            # ts['H'+str(enter_row)] = vf_0
            # ts['I'+str(enter_row)] = vf_0*days
            # ts['J'+str(enter_row)] = 2*vf_0*days
            ts['K'+str(enter_row)] = lkup_taxes[(lkup_taxes['SHIP']==ship)&(lkup_taxes['SAIL_FROM']==saildatefrom)]['GVT_TAX'].values[0]

            # set new enter_row
            enter_row += 1
        ts.title = sheet_name
        tb.save("temp_"+sheet_name)

        # # Executive Summary
        # es.title = sheet_name
        # eb.save("exec_"+sheet_name)


        st.success("KACHOW we're done! - - - Download offer sheets below.")

        for file in os.listdir("./"):
            if file.startswith("temp"):
                with open(file, 'rb') as my_file: #TODO: need to check for loop logic for wkbk to sheet, ships with offers sheet to offers with ships, single sheet each... need to store this info somehow
                    st.download_button(
                        label = 'Download '+file[5::],
                        data = my_file,
                        file_name = file[5::]+'_'+str(date.today())+'.xlsx',
                        mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         on_click = balloons()
                        )
            # if file.startswith("exec"):
            #     with open(file, 'rb') as my_file: #TODO: need to check for loop logic for wkbk to sheet, ships with offers sheet to offers with ships, single sheet each... need to store this info somehow
            #         st.download_button(
            #             label = 'Download Exec Summary',
            #             data = my_file,
            #             file_name = 'ExecutiveOfferSummary_'+str(date.today())+'.xlsx',
            #             mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            #             # on_click = balloons()
            #             )



        if balloons:
            st.balloons()
